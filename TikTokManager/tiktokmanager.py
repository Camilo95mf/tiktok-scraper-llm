"""
Class to manage data extraction from tiktok api. 
- code based on pyktok library
"""
import os
import asyncio
import shutil
from datetime import datetime
import time
import random
import json
import traceback

import pyktok as pyk
import requests
import pandas as pd

from TikTokApi import TikTokApi
# from TikTokApi.tiktok import EmptyResponseException
from openpyxl import load_workbook

ms_token = os.environ.get("ms_token", None)  # set your own ms_token
context_dict = {'viewport': {'width': 0,
                             'height': 0},
                'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36'}

class TikTokManager():
    """Class to mining data from tiktok api"""
    def __init__(self, output_path:str, temp_path:str) -> None:
        self.output_path = output_path
        self.temp_path = temp_path
        pyk.specify_browser('firefox')
        print(f'using ms_token: {ms_token}')

    def clear_folder(self, folder_path:str) -> None:
        """
        Clear or create a folder (used for temp and output folder)
        
        args:
        - folder_path: path to folder

        return:
        - None
        """
        # Create the folder if it doesn't exist
        print(f"Cleaning foder {folder_path}")
        os.makedirs(folder_path, exist_ok=True)

        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            # If it's a file or a symbolic link, delete it
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            # If it's a folder, delete it and all its contents
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)

    def save_to_excel(self, df: pd.DataFrame, file_path: str, sheet_name: str = "Outputs") -> None:
        """
        Append a DataFrame to an existing Excel file or create a new file if it doesn't exist.

        args:
        - df: DataFrame to append to the Excel file
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to append the DataFrame

        return:
        - None
        """
        if os.path.exists(file_path):
            # Load the existing workbook
            book = load_workbook(file_path)
            # Check if the sheet already exists
            if sheet_name in book.sheetnames:
                # Read the existing sheet into a DataFrame
                existing_df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
                # Concatenate the existing DataFrame with the new DataFrame
                updated_df = pd.concat([existing_df, df], ignore_index=True)
            else:
                updated_df = df
            
            # Write the updated DataFrame to the Excel file
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists='replace') as writer:
                updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # If the file doesn't exist, create a new one
            df.to_excel(file_path, sheet_name=sheet_name, index=False)

    async def get_video_urls(self,
                            tt_ent,
                            video_ct:int,
                            ent_type:str,
                            headless=True
                            ) -> list[str]:
        """
        Extract video urls based on tt_ent argument.

        args:
        - tt_ent: tiktok entity to extract videos
        - video_ct: amount of videos to extract
        - ent_type: type of entity to extract videos: "user", "hashtag", or "video_related"
        - headless: boolean to set headless mode on browser

        return:
        - List of video urls
        """
        if ent_type not in ['user','hashtag','video_related']:
            raise ValueError('Only allowed `ent_type` values are "user", "hashtag", or "video_related".')

        url_p1 = "https://www.tiktok.com/@"
        url_p2 = "/video/"
        tt_list = []

        async with TikTokApi() as api:
            await api.create_sessions(headless=headless,
                                    ms_tokens=[ms_token],
                                    num_sessions=1,
                                    sleep_after=40,
                                    context_options=context_dict)
            if ent_type == 'user':
                ent = api.user(tt_ent)
            elif ent_type == 'hashtag':
                ent = api.hashtag(name=tt_ent)
            elif ent_type == 'key_words':
                ent = api.search()
                ent.search_type(tt_ent, 'video', count=video_ct)
            else:
                ent = api.video(url=tt_ent)

            if ent_type in ['user','hashtag']:
                async for video in ent.videos(count=video_ct):
                    tt_list.append(video.as_dict)
            else:
                async for related_video in ent.related_videos(count=video_ct):
                    tt_list.append(related_video.as_dict)

        id_list = [i['id'] for i in tt_list]
        if ent_type == 'user':
            video_list = [url_p1 + tt_ent + url_p2 + i for i in id_list]
        else:
            author_list = [i['author']['uniqueId'] for i in tt_list]
            video_list = []
            for n, i in enumerate(author_list):
                video_url = url_p1 + author_list[n] + url_p2 + id_list[n]
                video_list.append(video_url)
        return video_list

    async def get_video_urls_v2(self,
                            tt_ent,
                            video_ct:int,
                            ent_type:str,
                            headless=True
                            ) -> list[str]:
        """
        Extract video urls based on tt_ent argument.

        args:
        - tt_ent: tiktok entity to extract videos
        - video_ct: amount of videos to extract
        - ent_type: type of entity to extract videos: "user", "hashtag", or "video_related"
        - headless: boolean to set headless mode on browser

        return:
        - List of video urls
        """
        if ent_type not in ['user','hashtag','video_related']:
            raise ValueError('Only allowed `ent_type` values are "user", "hashtag", or "video_related".')

        url_p1 = "https://www.tiktok.com/@"
        url_p2 = "/video/"
        tt_list = []
        end_flag = False
        retries = 0

        while retries < 5 and not end_flag:
            try:
                async with TikTokApi() as api:
                    await api.create_sessions(headless=headless,
                                            ms_tokens=[ms_token],
                                            num_sessions=1,
                                            sleep_after=10,
                                            context_options=context_dict,
                                            browser='chromium'
                                        )
                    if ent_type == 'user':
                        ent = api.user(tt_ent)
                    elif ent_type == 'hashtag':
                        ent = api.hashtag(name=tt_ent)
                    elif ent_type == 'key_words':
                        ent = api.search()
                        ent.search_type(tt_ent, 'video', count=video_ct)
                    else:
                        ent = api.video(url=tt_ent)

                    if ent_type in ['user','hashtag']:
                        async for video in ent.videos(count=video_ct):
                            time.sleep(random.randint(1, 2))
                            tt_list.append(video.as_dict)
                            if len(tt_list) >= video_ct:
                                end_flag = True
                                break
                    else:
                        async for related_video in ent.related_videos(count=video_ct):
                            tt_list.append(related_video.as_dict)
            except Exception as e:
                print(f"\n Error trying to mining videos for hashtag {tt_ent}: {str(e)} \n")
                print("Retrying...")
                del api
                retries += 1
                time.sleep(random.randint(1, 8))

        id_list = [i['id'] for i in tt_list]
        if ent_type == 'user':
            video_list = [url_p1 + tt_ent + url_p2 + i for i in id_list]
        else:
            author_list = [i['author']['uniqueId'] for i in tt_list]
            video_list = []
            for n, i in enumerate(author_list):
                video_url = url_p1 + author_list[n] + url_p2 + id_list[n]
                video_list.append(video_url)
        return video_list
    
    async def get_comments(self, video_id:str, comment_amount:int) -> list[dict]:
        """
        Extract comments from video using video id

        args:
        - video_id: video id
        - comment_amount: amount of comments to extract

        return:
        - List of dictionaries with comments info
        """
        count = 0
        retries = 0
        comment_list = []
        while retries < 5:
            await asyncio.sleep(1)
            async with TikTokApi() as api:
                video = api.video(id=video_id)
                await api.create_sessions(ms_tokens=[ms_token],
                                        num_sessions=1,
                                        sleep_after=5,
                                        context_options=context_dict
                                    )
                try:
                    async for comment in video.comments(count=comment_amount):
                        if count >= comment_amount:
                            break
                        aux_dict = comment.as_dict
                        dt_object = datetime.fromtimestamp(aux_dict.get("create_time"))
                        formated_comment = {
                            "video_id": video_id,
                            "language": aux_dict.get("comment_language"),
                            "text": aux_dict.get("text"),
                            "likes": aux_dict.get("digg_count"),
                            "date": dt_object.strftime('%d-%m-%Y %H:%M')
                        }
                        comment_list.append(formated_comment)
                        count += 1
                    return comment_list
                except Exception as e:
                    retries += 1
                    print(f"\n Error trying to get comments for video id {video_id}: {str(e)} \n")
                    print("Retrying...")
                    del video
                    time.sleep(random.randint(1, 8))

        return comment_list
    
    async def get_comments_v2(self, video_id:str, comment_amount:int) -> list[dict]:
        """
        Extract comments from video using video id

        args:
        - video_id: video id
        - comment_amount: amount of comments to extract

        return:
        - List of dictionaries with comments info
        """
        retries = 1
        comment_list = []
        end_flag = False
        while retries < 3 and not end_flag:
            try:
                async with TikTokApi() as api:
                    await api.create_sessions(ms_tokens=[ms_token],
                                            num_sessions=1,
                                            sleep_after=20,
                                            context_options=context_dict,
                                            browser='chromium',
                                            headless=False
                                        )
                    video = api.video(id=video_id)
                    
                    async for comment in video.comments(count=comment_amount):
                        time.sleep(random.randint(1, 5))
                        aux_dict = comment.as_dict
                        dt_object = datetime.fromtimestamp(aux_dict.get("create_time"))
                        formated_comment = {
                            "video_id": video_id,
                            "language": aux_dict.get("comment_language"),
                            "text": aux_dict.get("text"),
                            "likes": aux_dict.get("digg_count"),
                            "date": dt_object.strftime('%d-%m-%Y %H:%M')
                        }
                        comment_list.append(formated_comment)
                        if len(comment_list) >= comment_amount:
                            end_flag = True
                            break
                    return comment_list
            except Exception as e:
                retries += 1
                print(f"\n Error trying to get comments for video id {video_id}: {str(e)} \n")
                traceback.print_exc()
                print("Retrying...")
                del video
                time.sleep(random.randint(1, 8))

        return comment_list
        
    def extract_videos_data(self, hashtag_list:list[str], video_amount:int, comment_amount:int) -> dict:
        """
        Extract relevant info from videos related with hashtag parameter
        
        args:
        - hashtag_list: list of hashtags to extract videos
        - video_amount: amount of videos to extract per hashtag
        - comment_amount: amount of comments to extract per video
        
        return:
        - Dictionary with the path to the excel file with the extracted data
        """
        tiktok_url = 'https://www.tiktok.com/@tiktok/video/'
        temp_data_path = f'{self.temp_path}/temp_data.csv'

        now = datetime.now()
        formatted_datetime = now.strftime("%m-%d-%Y_%H%M")
        output_path = f'{self.output_path}/video_info_{formatted_datetime}.xlsx'

        self.clear_folder(self.temp_path)

        url_list = []
        url_per_hashtag = {}


        for hashtag in hashtag_list:
            print(f"\n\n\nMining {video_amount} video data for keyword: {hashtag}")
            url_list_extracted = asyncio.run(self.get_video_urls(
                                                tt_ent=hashtag,
                                                ent_type="hashtag",
                                                video_ct=video_amount,
                                            ))
            
            set1 = set(url_list)
            set1.update([url for url in url_list_extracted if url not in set1])
            url_list = list(set1)

            
            for url in url_list:
                time.sleep(random.randint(1, 5))
                try:
                    pyk.save_tiktok(url,False,temp_data_path)
                except Exception as e:
                    print(f"Error saving video data from url {url}: {str(e)} \n\n")
                    time.sleep(random.randint(1, 3))
                    continue
            
            url_per_hashtag[hashtag] = url_list
        
        data_df = pd.read_csv(temp_data_path, sep=',', dtype=str)
        data_df['hashtag'] = ''
        data_df['trasncription_lang'] = ''
        data_df['video_trasncription'] = ''
        data_df['mining_date'] = formatted_datetime
        comment_list = []

        for index, row in data_df.iterrows():
            # Get transcription of video
            print("\nMining transcription from video: ",row['video_id'])
            aux_url = tiktok_url+row['video_id']
            tiktok_json = pyk.alt_get_tiktok_json(aux_url)
            if tiktok_json is not None:

                trasncription_list = tiktok_json.get("__DEFAULT_SCOPE__").get("webapp.video-detail").get("itemInfo").get("itemStruct").get("video").get("subtitleInfos")
                if len(trasncription_list) > 0:
                    # Filter transcription by english language
                    eng_transcription = [d for d in trasncription_list if "eng-" in d.get("LanguageCodeName", "")]
                    if len(eng_transcription) > 0:
                        language = eng_transcription[0].get("LanguageCodeName")
                        trasncription_url = eng_transcription[0].get("Url")
                    else:
                        language = trasncription_list[0].get("LanguageCodeName")
                        trasncription_url = trasncription_list[0].get("Url")
                    response = requests.get(trasncription_url, timeout=15)
                    result_array = response.text.split('\n')
                    result_array = [item for item in result_array if item and "-->" not in item and "WEBVTT" not in item]
                    final_trasncription =' '.join(result_array)

                    data_df.at[index, 'trasncription_lang'] = language
                    data_df.at[index, 'video_trasncription'] = final_trasncription

                    for key, url_list in url_per_hashtag.items():
                        if any(data_df.at[index, 'video_id'] in string for string in url_list):
                            data_df.at[index, 'hashtag'] = key
                        
            # Get comments for each video
            print("Mining comments from video: ",row['video_id'])
            comment_list += asyncio.run(self.get_comments(row['video_id'], comment_amount))
        
        comment_df = pd.DataFrame(comment_list)
        
        self.save_to_excel(data_df, output_path, 'videos_data')
        self.save_to_excel(comment_df, output_path, 'comments_data')

        return {"video_data_path": output_path}
    
    def extract_videos_data_v2(self, hashtag_list:list[str], video_amount:int, comment_amount:int) -> dict:
        """
        Extract relevant info from videos related with hashtag parameter
        
        args:
        - hashtag_list: list of hashtags to extract videos
        - video_amount: amount of videos to extract per hashtag
        - comment_amount: amount of comments to extract per video
        
        return:
        - Dictionary with the path to the excel file with the extracted data
        """

        now = datetime.now()
        formatted_datetime = now.strftime("%m-%d-%Y_%H%M")
        output_path = f'{self.output_path}/video_info_{formatted_datetime}.xlsx'

        # self.clear_folder(self.temp_path)

        url_list = []
        url_per_hashtag = {}


        for hashtag in hashtag_list:
            print(f"\n\n\nMining {video_amount} video(s) for keyword: {hashtag}")
            url_list_extracted = asyncio.run(self.get_video_urls_v2(
                                                tt_ent=hashtag,
                                                ent_type="hashtag",
                                                video_ct=video_amount,
                                            ))
            
            set1 = set(url_list)
            set1.update([url for url in url_list_extracted if url not in set1])
            url_list = list(set1)
            url_per_hashtag[hashtag] = url_list

            print(f'Got {len(url_list)} urls for hashtag: {hashtag}')

        seen = set()
        url_per_hashtag_validated = {}
        # Remove duplicates from url_per_hashtag
        for key, value_list in url_per_hashtag.items():
            unique_items = []
            for item in value_list:
                if item not in seen:
                    seen.add(item)
                    unique_items.append(item)     
            url_per_hashtag_validated[key] = unique_items
        
        with open(f'{self.temp_path}/video_urls_checkpoint_{formatted_datetime}.json', 'w', encoding='utf-8') as file:
            print(f"Saving video urls checkpoint in {self.temp_path}")
            file.write(json.dumps(url_per_hashtag_validated, indent=4))
        
        for hashtag, url_arr in url_per_hashtag_validated.items():
            data_df = None
            comment_df = None
            comment_list = []
            data_list = []
            for url in url_arr:
                print(f"\nMining data from video: {url}")
                row_dict = {}
                tiktok_json = pyk.alt_get_tiktok_json(url)

                # with open(f'{self.temp_path}/test_json.json', 'w', encoding='utf-8') as file:
                #     file.write(json.dumps(tiktok_json))

                if tiktok_json is not None:
                    try:
                        data_slot = tiktok_json["__DEFAULT_SCOPE__"]['webapp.video-detail']['itemInfo']['itemStruct']
                        video_id = data_slot['id']
                        row_dict['video_id'] = video_id
                        row_dict['hashtag'] = hashtag
                        created_time = data_slot.get('createTime')
                        if created_time is not None:
                            row_dict['video_timestamp'] = datetime.fromtimestamp(int(created_time)).isoformat()
                        else:
                            row_dict['video_timestamp'] = None
                        row_dict['video_locationcreated'] = data_slot.get('locationCreated')
                        row_dict['video_diggcount'] = data_slot.get('stats').get('diggCount')
                        row_dict['video_sharecount'] = data_slot.get('stats').get('shareCount')
                        row_dict['video_commentcount'] = data_slot.get('stats').get('commentCount')
                        row_dict['video_playcount'] = data_slot.get('stats').get('playCount')
                        row_dict['video_description'] = data_slot.get('desc')
                        row_dict['video_is_ad'] = data_slot.get('isAd')               
                        row_dict['author_username'] = data_slot.get('author').get('uniqueId')
                        row_dict['author_name'] = data_slot.get('author').get('nickname')
                        row_dict['author_followercount'] = data_slot.get('authorStats').get('followerCount')
                        
                        
                        row_dict['author_heartcount'] = data_slot.get('authorStats').get('heartCount')
                        row_dict['author_videocount'] = data_slot.get('authorStats').get('videoCount')
                        row_dict['author_diggcount'] = data_slot.get('authorStats').get('diggCount')
                        row_dict['author_verified'] = data_slot.get('author').get('verified')


                        trasncription_list = data_slot.get("video").get("subtitleInfos")
                        if len(trasncription_list) > 0:
                            print(f"Mining transcription from video: {url}")
                            # Filter transcription by english language
                            eng_transcription = [d for d in trasncription_list if "eng-" in d.get("LanguageCodeName", "")]
                            if len(eng_transcription) > 0:
                                language = eng_transcription[0].get("LanguageCodeName")
                                trasncription_url = eng_transcription[0].get("Url")
                            else:
                                language = trasncription_list[0].get("LanguageCodeName")
                                trasncription_url = trasncription_list[0].get("Url")
                            response = requests.get(trasncription_url, timeout=15)
                            result_array = response.text.split('\n')
                            result_array = [item for item in result_array if item and "-->" not in item and "WEBVTT" not in item]
                            final_trasncription =' '.join(result_array)

                            row_dict['trasncription_lang'] = language
                            row_dict['video_trasncription'] = final_trasncription
                        else:
                            print(f"Error empty trasncription data from video: {url}")
                        
                        row_dict['mining_date'] = formatted_datetime

                        # data_list.append(row_dict)
                        data_list = [row_dict]
                        data_df = pd.DataFrame(data_list)
                        self.save_to_excel(data_df, output_path, 'videos_data')
                            
                        # Get comments for each video
                        print(f"Mining comments from video: {video_id}")
                        # comment_list += asyncio.run(self.get_comments(video_id, comment_amount))
                        comment_list = asyncio.run(self.get_comments(video_id, comment_amount))
                        comment_df = pd.DataFrame(comment_list)
                        self.save_to_excel(comment_df, output_path, 'comments_data')
                    except Exception as e:
                        print(f"Error saving video data from url {url}: {str(e)} \n\n")
                        time.sleep(random.randint(1, 2))
                        continue
                else:
                    print(f"Error empty json data from video: {url}")
        
            # comment_df = pd.DataFrame(comment_list)
            # self.save_to_excel(comment_df, output_path, 'comments_data')

            # data_df = pd.DataFrame(data_list)
            # self.save_to_excel(data_df, output_path, 'videos_data')

            del data_df
            del comment_df

        return {"video_data_path": output_path}
    
    def get_video_transcription(self, videos_id:list[str]):
        """
        Get video transcription from a list of video ids and save it in a txt file

        args:
        - videos_id: list of video ids

        return:
        - temp_data_path: path to the txt file with the transcriptions
        """
        tiktok_url = 'https://www.tiktok.com/@tiktok/video/'
        temp_data_path = f'{self.temp_path}/transcriptions.txt'
        res_list = []
        for id_v in videos_id:
            # Get transcription of video
            aux_url = tiktok_url+id_v
            tiktok_json = pyk.alt_get_tiktok_json(aux_url)
            if tiktok_json is not None:
                trasncription_list = tiktok_json.get("__DEFAULT_SCOPE__").get("webapp.video-detail").get("itemInfo").get("itemStruct").get("video").get("subtitleInfos")
                if len(trasncription_list) > 0:
                    # language = trasncription_list[0].get("LanguageCodeName")
                    trasncription_url = trasncription_list[0].get("Url")
                    response = requests.get(trasncription_url, timeout=6)
                    result_array = response.text.split('\n')
                    result_array = [item for item in result_array if item and "-->" not in item and "WEBVTT" not in item]
                    final_trasncription =' '.join(result_array)
                    res_list.append(id_v+":\n"+final_trasncription+"\n\n")
        with open(temp_data_path, "w", encoding="utf-8") as file:
            # Write each string in the list to the file
            for item in res_list:
                file.write(item + "\n")

        return temp_data_path
        
